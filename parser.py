"""
Parser for Russian financial reports (БФО) from bo.nalog.gov.ru.
Extracts balance sheet and P&L data from xlsx files.
"""

import re
import openpyxl


# Row codes for Balance Sheet
BALANCE_CODES = {
    "1100": "Внеоборотные активы",
    "1110": "Нематериальные активы",
    "1150": "Основные средства",
    "1170": "Финансовые вложения (долгоср.)",
    "1210": "Запасы",
    "1220": "НДС",
    "1230": "Дебиторская задолженность",
    "1240": "Фин. вложения (краткоср.)",
    "1250": "Денежные средства",
    "1260": "Прочие оборотные активы",
    "1200": "Оборотные активы",
    "1600": "БАЛАНС (актив)",
    "1300": "Капитал и резервы",
    "1410": "Долгосрочные заёмные средства",
    "1510": "Краткосрочные заёмные средства",
    "1520": "Кредиторская задолженность",
    "1700": "БАЛАНС (пассив)",
}

# Row codes for P&L
PL_CODES = {
    "2110": "Выручка",
    "2120": "Себестоимость продаж",
    "2100": "Валовая прибыль",
    "2210": "Коммерческие расходы",
    "2220": "Управленческие расходы",
    "2200": "Прибыль от продаж",
    "2310": "Доходы от участия",
    "2320": "Проценты к получению",
    "2330": "Проценты к уплате",
    "2340": "Прочие доходы",
    "2350": "Прочие расходы",
    "2300": "Прибыль до налогообложения",
    "2410": "Налог на прибыль",
    "2400": "Чистая прибыль",
}

# Known company name mappings (short code -> full name, INN)
COMPANY_MAP = {
    "ЕЯ": ("ООО \"Екатеринбург Яблоко\"", "6670381056"),
    "ЛИ": ("ООО \"ЛАБ ИНДАСТРИЗ\"", "7702691545"),
    "ГРАДИЕНТ": ("ООО \"НТС \"Градиент\"", "7720125736"),
    "НК": ("АО \"НЭФИС КОСМЕТИКС\"", "1653005126"),
    "СИНЕРГЕТИК": ("ООО \"СИНЕРГЕТИК\"", "5257123941"),
    "ЮНИК": ("ООО \"ЮНИКОСМЕТИК\"", "7826704356"),
    "АРНЕСТ": ("АО \"АРНЕСТ\"", "2631006752"),
    "МИКСИТ": ("ООО \"МИКСИТ\"", "7733333130"),
    "СПЛАТ_Г": ("ООО \"СПЛАТ ГЛОБАЛ\"", "7718173605"),
    "СПЛАТ": ("ООО \"СПЛАТ\"", "7703539871"),
    "ГЕЛЬТЕК": ("ООО \"ГЕЛЬТЕК\"", "5017127741"),
    "ЗЕТТЕК": ("ООО \"ЗетТек\"", "7701848183"),
    "АН": ("ООО \"Аэрозоль Новомосковск\"", "7116010113"),
    "БИГ": ("ООО \"БИГ\"", "7751057596"),
    "СИБИАР": ("АО \"СИБИАР\"", "5404105343"),
    "ГРАНД": ("ООО \"ГРАНД А.В.\"", "7703528781"),
    "СХЗ": ("АО \"СТУПИНСКИЙ ХИМИЧЕСКИЙ ЗАВОД\"", "5045022211"),
    "ФЛ": ("АО \"ФАБЕРЛИК\"", "5001026970"),
    "ШЕШМЕШ": ("ООО \"ШЕМЕШ\"", "7736347445"),
    "ЭКТ": ("ООО \"ЭКТ\"", "9704009302"),
}


def parse_value(val):
    """Parse a financial value like '1 234 567' or '(1 234)' into a number."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ('-', '', '(-)', 'н/д', '0'):
        return 0.0
    negative = False
    if s.startswith('(') and s.endswith(')'):
        negative = True
        s = s[1:-1]
    s = s.replace('\xa0', '').replace(' ', '').replace(',', '.')
    s = re.sub(r'[^\d.\-]', '', s)
    if not s or s == '-':
        return 0.0
    try:
        result = float(s)
        return -result if negative else result
    except ValueError:
        return None


def extract_sheet_data(ws, code_map):
    """Extract data from a sheet using row codes."""
    code_col = None
    code_row_start = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or "").strip()
            if v == "Код строки":
                code_col = c
                code_row_start = r + 1
                break
            if v == "3" and c > 3:
                for cr in range(r + 1, min(r + 10, ws.max_row + 1)):
                    cv = str(ws.cell(row=cr, column=c).value or "").strip()
                    if re.match(r'^\d{4}$', cv):
                        code_col = c
                        code_row_start = r + 1
                        break
                if code_col:
                    break
        if code_col:
            break

    if code_col is None:
        all_codes = set(code_map.keys())
        for r in range(5, min(25, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = str(ws.cell(row=r, column=c).value or "").strip()
                if v in all_codes:
                    code_col = c
                    code_row_start = r
                    break
            if code_col:
                break

    if code_col is None:
        return None, None

    val_cols = []
    for r in range(code_row_start or 6, min((code_row_start or 6) + 15, ws.max_row + 1)):
        code_val = str(ws.cell(row=r, column=code_col).value or "").strip()
        if code_val not in code_map:
            continue
        for c in range(code_col + 1, ws.max_column + 1):
            cv = ws.cell(row=r, column=c).value
            if cv is not None:
                sv = str(cv).strip()
                if sv and (re.search(r'\d', sv) or sv == '-'):
                    val_cols.append(c)
        break

    if not val_cols:
        val_cols = [code_col + 2, code_col + 5]
        val_cols = [c for c in val_cols if c <= ws.max_column]

    current_data = {}
    prev_data = {}
    for r in range(code_row_start or 6, ws.max_row + 1):
        code_val = str(ws.cell(row=r, column=code_col).value or "").strip()
        if code_val in code_map:
            if len(val_cols) >= 1:
                current_data[code_val] = parse_value(ws.cell(row=r, column=val_cols[0]).value)
            if len(val_cols) >= 2:
                prev_data[code_val] = parse_value(ws.cell(row=r, column=val_cols[1]).value)

    return current_data, prev_data


def parse_xlsx(filepath):
    """Parse a single xlsx report file. Returns dict with balance and pl data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"balance": {}, "pl": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lower = sheet_name.lower()
        if "баланс" in sheet_lower:
            curr, _ = extract_sheet_data(ws, BALANCE_CODES)
            if curr:
                result["balance"] = curr
        elif "результат" in sheet_lower or "финанс" in sheet_lower:
            curr, _ = extract_sheet_data(ws, PL_CODES)
            if curr:
                result["pl"] = curr

    wb.close()
    return result


def detect_company_year(filename):
    """Detect company short code and year from filename like 'АРНЕСТ_2024.xlsx'."""
    stem = filename.rsplit('.', 1)[0] if '.' in filename else filename
    parts = stem.rsplit('_', 1)
    if len(parts) != 2:
        return None, None
    short_name, year_str = parts
    try:
        year = int(year_str)
    except ValueError:
        return None, None
    if year < 2015 or year > 2030:
        return None, None
    return short_name.strip(), year


def compute_metrics(pl, bal):
    """Compute derived financial metrics from raw P&L and balance data."""
    revenue = pl.get("2110", 0) or 0
    cogs = abs(pl.get("2120", 0) or 0)
    gross_profit = pl.get("2100", 0) or 0
    operating_profit = pl.get("2200", 0) or 0
    net_profit = pl.get("2400", 0) or 0
    commercial = abs(pl.get("2210", 0) or 0)
    admin = abs(pl.get("2220", 0) or 0)
    interest_income = pl.get("2320", 0) or 0
    interest_expense = abs(pl.get("2330", 0) or 0)
    other_income = pl.get("2340", 0) or 0
    other_expense = abs(pl.get("2350", 0) or 0)
    profit_before_tax = pl.get("2300", 0) or 0

    total_assets = bal.get("1600", 0) or bal.get("1700", 0) or 0
    equity = bal.get("1300", 0) or 0
    non_current = bal.get("1100", 0) or 0
    current_assets = bal.get("1200", 0) or 0
    inventories = bal.get("1210", 0) or 0
    receivables = bal.get("1230", 0) or 0
    cash = bal.get("1250", 0) or 0
    fixed_assets = bal.get("1150", 0) or 0
    lt_debt = bal.get("1410", 0) or 0
    st_debt = bal.get("1510", 0) or 0
    payables = bal.get("1520", 0) or 0

    gross_margin = (gross_profit / revenue * 100) if revenue else None
    operating_margin = (operating_profit / revenue * 100) if revenue else None
    net_margin = (net_profit / revenue * 100) if revenue else None
    roa = (net_profit / total_assets * 100) if total_assets else None
    roe = (net_profit / equity * 100) if equity and equity > 0 else None
    total_debt = lt_debt + st_debt
    debt_equity = (total_debt / equity) if equity and equity > 0 else None
    current_liabilities = st_debt + payables
    current_ratio = (current_assets / current_liabilities) if current_liabilities else None

    return {
        "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
        "operating_profit": operating_profit, "net_profit": net_profit,
        "commercial_expenses": commercial, "admin_expenses": admin,
        "interest_income": interest_income, "interest_expense": interest_expense,
        "other_income": other_income, "other_expense": other_expense,
        "profit_before_tax": profit_before_tax,
        "total_assets": total_assets, "equity": equity,
        "non_current_assets": non_current, "current_assets": current_assets,
        "inventories": inventories, "receivables": receivables,
        "cash": cash, "fixed_assets": fixed_assets,
        "lt_debt": lt_debt, "st_debt": st_debt, "payables": payables,
        "gross_margin": round(gross_margin, 2) if gross_margin is not None else None,
        "operating_margin": round(operating_margin, 2) if operating_margin is not None else None,
        "net_margin": round(net_margin, 2) if net_margin is not None else None,
        "roa": round(roa, 2) if roa is not None else None,
        "roe": round(roe, 2) if roe is not None else None,
        "debt_equity": round(debt_equity, 2) if debt_equity is not None else None,
        "current_ratio": round(current_ratio, 2) if current_ratio is not None else None,
    }


def get_company_info(short_code):
    """Get full name and INN for a company short code."""
    if short_code in COMPANY_MAP:
        name, inn = COMPANY_MAP[short_code]
        return name, inn
    return short_code, ""

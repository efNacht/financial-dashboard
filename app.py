"""
Financial Dashboard — Flask web application.
Parses Russian financial reports (БФО) and displays interactive analytics.
"""

import os
import json
import sqlite3
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from parser import parse_xlsx, detect_company_year, compute_metrics, get_company_info

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

APP_DIR = Path(__file__).resolve().parent
DB_PATH = os.environ.get('DB_PATH', str(APP_DIR / 'data.db'))
UPLOAD_DIR = Path(os.environ.get('UPLOAD_DIR', str(APP_DIR / 'uploads')))
UPLOAD_DIR.mkdir(exist_ok=True)


# ============================================================
# DATABASE
# ============================================================

def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    return db


def init_db():
    db = get_db()
    db.execute('''CREATE TABLE IF NOT EXISTS companies (
        short_code TEXT PRIMARY KEY,
        full_name TEXT NOT NULL,
        inn TEXT DEFAULT ''
    )''')
    db.execute('''CREATE TABLE IF NOT EXISTS reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        short_code TEXT NOT NULL,
        year INTEGER NOT NULL,
        metrics TEXT NOT NULL,
        raw_pl TEXT DEFAULT '{}',
        raw_balance TEXT DEFAULT '{}',
        filename TEXT DEFAULT '',
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(short_code, year),
        FOREIGN KEY (short_code) REFERENCES companies(short_code)
    )''')
    db.commit()
    db.close()


def get_all_data():
    """Get all company data formatted for the dashboard JS."""
    db = get_db()
    companies = {r['short_code']: {'full_name': r['full_name'], 'inn': r['inn']}
                 for r in db.execute('SELECT * FROM companies').fetchall()}

    reports = db.execute('SELECT * FROM reports ORDER BY short_code, year').fetchall()
    db.close()

    result = {}
    for r in reports:
        sc = r['short_code']
        if sc not in result:
            info = companies.get(sc, {'full_name': sc, 'inn': ''})
            result[sc] = {
                'full_name': info['full_name'],
                'inn': info['inn'],
                'short_name': sc,
                'years_data': {},
            }
        result[sc]['years_data'][r['year']] = json.loads(r['metrics'])

    # Sort by latest revenue descending
    def sort_key(item):
        years = item[1]['years_data']
        if not years:
            return 0
        latest = max(years.keys())
        return -(years[latest].get('revenue', 0) or 0)

    result = dict(sorted(result.items(), key=sort_key))
    return result


def save_report(short_code, year, metrics, raw_pl, raw_balance, filename):
    """Save or update a parsed report in the database."""
    full_name, inn = get_company_info(short_code)
    db = get_db()
    db.execute('INSERT OR REPLACE INTO companies (short_code, full_name, inn) VALUES (?, ?, ?)',
               (short_code, full_name, inn))
    db.execute('''INSERT OR REPLACE INTO reports (short_code, year, metrics, raw_pl, raw_balance, filename)
                  VALUES (?, ?, ?, ?, ?, ?)''',
               (short_code, year, json.dumps(metrics, ensure_ascii=False),
                json.dumps(raw_pl, ensure_ascii=False), json.dumps(raw_balance, ensure_ascii=False),
                filename))
    db.commit()
    db.close()


# ============================================================
# ROUTES
# ============================================================

@app.route('/')
def index():
    data = get_all_data()
    return render_template('dashboard.html', data_json=json.dumps(data, ensure_ascii=False))


@app.route('/dashboard')
def dashboard():
    data = get_all_data()
    return render_template('dashboard.html', data_json=json.dumps(data, ensure_ascii=False))


@app.route('/summary')
def summary():
    data = get_all_data()
    return render_template('summary.html', data_json=json.dumps(data, ensure_ascii=False))


@app.route('/upload', methods=['GET'])
def upload_page():
    db = get_db()
    reports = db.execute('''SELECT c.full_name, r.short_code, r.year, r.filename, r.uploaded_at
                           FROM reports r JOIN companies c ON r.short_code = c.short_code
                           ORDER BY r.uploaded_at DESC''').fetchall()
    companies = db.execute('SELECT short_code, full_name FROM companies ORDER BY full_name').fetchall()
    db.close()
    return render_template('upload.html', reports=reports, companies=companies)


@app.route('/api/upload', methods=['POST'])
def api_upload():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'Файлы не выбраны'}), 400

    results = []
    for f in files:
        if not f.filename or not f.filename.endswith('.xlsx'):
            results.append({'file': f.filename, 'status': 'error', 'message': 'Не xlsx файл'})
            continue

        short_code, year = detect_company_year(f.filename)
        if not short_code or not year:
            results.append({'file': f.filename, 'status': 'error',
                           'message': 'Не удалось определить компанию/год из имени файла. Формат: КОМПАНИЯ_ГОД.xlsx'})
            continue

        # Save file temporarily
        filepath = UPLOAD_DIR / f.filename
        f.save(filepath)

        try:
            raw = parse_xlsx(str(filepath))
            metrics = compute_metrics(raw.get('pl', {}), raw.get('balance', {}))
            save_report(short_code, year, metrics, raw.get('pl', {}), raw.get('balance', {}), f.filename)

            full_name, _ = get_company_info(short_code)
            results.append({
                'file': f.filename,
                'status': 'ok',
                'company': full_name,
                'short_code': short_code,
                'year': year,
                'revenue': metrics.get('revenue', 0),
                'net_profit': metrics.get('net_profit', 0),
            })
        except Exception as e:
            results.append({'file': f.filename, 'status': 'error', 'message': str(e)})
        finally:
            filepath.unlink(missing_ok=True)

    return jsonify({'results': results})


@app.route('/api/delete', methods=['POST'])
def api_delete():
    data = request.json
    short_code = data.get('short_code')
    year = data.get('year')
    db = get_db()
    if year:
        db.execute('DELETE FROM reports WHERE short_code = ? AND year = ?', (short_code, int(year)))
    else:
        db.execute('DELETE FROM reports WHERE short_code = ?', (short_code,))
        db.execute('DELETE FROM companies WHERE short_code = ?', (short_code,))
    db.commit()
    db.close()
    return jsonify({'ok': True})


@app.route('/api/data')
def api_data():
    return jsonify(get_all_data())


# Metric definitions shared with summary.html (key -> label, flags)
METRIC_DEFS = [
    {'key': 'revenue',             'label': 'Выручка',          'profit': False, 'is_pct': False},
    {'key': 'net_profit',          'label': 'Чистая прибыль',   'profit': True,  'is_pct': False},
    {'key': 'gross_profit',        'label': 'Валовая прибыль',  'profit': True,  'is_pct': False},
    {'key': 'operating_profit',    'label': 'Прибыль от продаж','profit': True,  'is_pct': False},
    {'key': 'profit_before_tax',   'label': 'Прибыль до налог.','profit': True,  'is_pct': False},
    {'key': 'cogs',                'label': 'Себестоимость',    'profit': False, 'is_pct': False},
    {'key': 'commercial_expenses', 'label': 'Коммерч. расходы', 'profit': False, 'is_pct': False},
    {'key': 'admin_expenses',      'label': 'Управл. расходы',  'profit': False, 'is_pct': False},
    {'key': 'gross_margin',        'label': 'Валовая рентаб.',  'profit': False, 'is_pct': True},
    {'key': 'operating_margin',    'label': 'Операц. рентаб.',  'profit': False, 'is_pct': True},
    {'key': 'net_margin',          'label': 'Чистая рентаб.',   'profit': False, 'is_pct': True},
    {'key': 'total_assets',        'label': 'Активы',           'profit': False, 'is_pct': False},
    {'key': 'equity',              'label': 'Собств. капитал',  'profit': False, 'is_pct': False},
]
METRIC_MAP = {m['key']: m for m in METRIC_DEFS}


@app.route('/export/xlsx')
def export_xlsx():
    """Export summary table as xlsx. Supports query params: year, metrics, hidden_types, hidden_companies."""
    data = get_all_data()

    # Determine current year
    all_years = set()
    for c in data.values():
        all_years.update(c['years_data'].keys())
    all_years = sorted(all_years)
    default_year = all_years[-1] if all_years else 2025

    try:
        current_year = int(request.args.get('year', default_year))
    except (ValueError, TypeError):
        current_year = default_year
    ym1 = current_year - 1
    ym2 = current_year - 2

    metrics_param = request.args.get('metrics', 'revenue,net_profit')
    active_keys = [k for k in metrics_param.split(',') if k in METRIC_MAP] or ['revenue', 'net_profit']
    active_metrics = [METRIC_MAP[k] for k in active_keys]

    hidden_types = set(request.args.get('hidden_types', '').split(',')) - {''}
    hidden_companies = set(request.args.get('hidden_companies', '').split(',')) - {''}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводная таблица"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Build headers and column metadata
    headers = ['Юр. лицо']
    col_meta = [{'type': 'name'}]
    for m in active_metrics:
        if 'ym2' not in hidden_types:
            headers.append(f"{m['label']} {ym2}"); col_meta.append({'type': 'val', 'metric': m, 'year': ym2})
        if 'ym1' not in hidden_types:
            headers.append(f"{m['label']} {ym1}"); col_meta.append({'type': 'val', 'metric': m, 'year': ym1})
        if 'y' not in hidden_types:
            headers.append(f"{m['label']} {current_year}"); col_meta.append({'type': 'val', 'metric': m, 'year': current_year})
        if 'pct' not in hidden_types:
            headers.append(f"{m['label']} {current_year}/{ym1}, %"); col_meta.append({'type': 'pct', 'metric': m})
        if 'diff' not in hidden_types:
            headers.append(f"{m['label']} {current_year}−{ym1}, руб."); col_meta.append({'type': 'diff', 'metric': m})

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_num = 2
    for sc, company in data.items():
        if sc in hidden_companies:
            continue
        yd = company['years_data']
        # Normalize keys to ints
        yd_n = {int(k): v for k, v in yd.items()}

        for col, meta in enumerate(col_meta, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

            if meta['type'] == 'name':
                cell.value = company['full_name']
                cell.alignment = Alignment(horizontal="left")
                continue

            m = meta['metric']
            v_y   = (yd_n.get(current_year) or {}).get(m['key'])
            v_ym1 = (yd_n.get(ym1) or {}).get(m['key'])

            if meta['type'] == 'val':
                v = (yd_n.get(meta['year']) or {}).get(m['key'])
                cell.value = v if v not in (0, None) else None
                if m['is_pct']:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
                if m['profit'] and isinstance(v, (int, float)) and v < 0:
                    cell.font = Font(color="DC2626")
            elif meta['type'] == 'pct':
                pct = (v_y - v_ym1) / abs(v_ym1) if (v_y is not None and v_ym1 not in (None, 0)) else None
                cell.value = pct
                cell.number_format = '0%;[Red]-0%'
                cell.alignment = Alignment(horizontal="center")
                if isinstance(pct, (int, float)) and pct < 0:
                    cell.font = Font(color="DC2626")
            elif meta['type'] == 'diff':
                diff = (v_y - v_ym1) if (v_y is not None and v_ym1 is not None) else None
                cell.value = diff
                cell.number_format = '#,##0;[Red]-#,##0'
                cell.alignment = Alignment(horizontal="right")
                if isinstance(diff, (int, float)) and diff < 0:
                    cell.font = Font(color="DC2626")

        row_num += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.column_dimensions['A'].width = 35

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=f"сводная_таблица_{current_year}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# SEED DATA — load existing xlsx files on first run
# ============================================================

def seed_from_directory(directory):
    """Parse all xlsx files from a directory and save to DB."""
    d = Path(directory)
    if not d.exists():
        return 0
    count = 0
    for fp in sorted(d.glob('*.xlsx')):
        short_code, year = detect_company_year(fp.name)
        if not short_code or not year:
            continue
        try:
            raw = parse_xlsx(str(fp))
            metrics = compute_metrics(raw.get('pl', {}), raw.get('balance', {}))
            save_report(short_code, year, metrics, raw.get('pl', {}), raw.get('balance', {}), fp.name)
            count += 1
        except Exception as e:
            print(f"  Error parsing {fp.name}: {e}")
    return count


# ============================================================
# INIT — runs on import (gunicorn, preview, or __main__)
# ============================================================

init_db()

# Seed from local directory if DB is empty
_db = get_db()
_cnt = _db.execute('SELECT COUNT(*) as c FROM reports').fetchone()['c']
_db.close()
if _cnt == 0:
    _reports_dir = os.environ.get('REPORTS_DIR', '')
    if _reports_dir and Path(_reports_dir).exists():
        print(f"Seeding from {_reports_dir}...")
        _n = seed_from_directory(_reports_dir)
        print(f"Loaded {_n} reports")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', '0') == '1')
